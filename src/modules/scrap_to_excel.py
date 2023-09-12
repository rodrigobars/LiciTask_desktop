import os
import pandas as pd
from time import sleep
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService

def extract_excel(path):       
    dados = pd.read_excel(path, header=None)
        
    for cel in dados.iloc[-1]:
        if type(cel)==str and "TOTAL" in cel:
            dados.drop(dados.index[-1], inplace=True)
            break
        
    dados.drop(0, inplace=True)
    header = dados.iloc[0].values
    dados = dados.iloc[1:,:]
    dados.columns = header
    dados.set_index(header[0], inplace=True)
    dados.reset_index(inplace=True)
    
    dados["DESCRIÇÃO / ESPECIFICAÇÃO"] = dados["DESCRIÇÃO / ESPECIFICAÇÃO"].astype(str)
    dados["VALOR DE REFERÊNCIA (unitário)(R$)"] = dados["VALOR DE REFERÊNCIA (unitário)(R$)"].astype(float).round(2)
    dados["VALOR DE REFERÊNCIA (total)(R$)"] = dados["VALOR DE REFERÊNCIA (total)(R$)"].astype(float).round(2)
    
    return dados

def scrap_to_excel(pregao, excel_path):
    BaseDado = extract_excel(excel_path)
    BaseDado = BaseDado.loc[:, :"VALOR DE REFERÊNCIA (unitário)(R$)"]

    # Criando colunas vazias no DataFrame
    col_names = ["ValorFornecedor", "Pos 1",
                "Pos 2", "Pos 3", "Pos 4", "Pos 5", "Status"]

    # ANOTAÇÃO: Necessito modificar isso
    for col in col_names:
        BaseDado.loc[0, col] = None

    # Armazenando o link para ser aberto no navegador posteriormente
    url = "http://comprasnet.gov.br/acesso.asp?url=/livre/Pregao/lista_pregao_filtro.asp?Opc=2"

    # Informando o Path para o arquivo "chromedriver.exe" e armazenando em "driver"
    op = webdriver.ChromeOptions()
    op.add_argument("--start-maximized")
    op.add_experimental_option("excludeSwitches", ["enable-logging"])

    # Format Path
    app_path = os.path.abspath(os.getcwd())
    
    # Chromebrowser path
    chrome_folder = "chrome-driver/chrome-win64/chrome.exe"
    chrome_path = os.path.join(app_path, chrome_folder)

    # Chromedriver path
    driver_folder = "chrome-driver/chromedriver-win64/chromedriver.exe"
    driver_path = os.path.join(app_path, driver_folder)

    # Informando o navegador que será utilizado
    op.binary_location = chrome_path
    driver = webdriver.Chrome(service=ChromeService(driver_path), options=op)

    # Abrindo o google chrome no site informado
    driver.get(url)

    # Navegador em tela cheia
    driver.maximize_window()

    # Configurando o tempo de espera em segundos até que o elemento html seja encontrado
    wdw = WebDriverWait(driver, 2000)
    wdw2 = WebDriverWait(driver, 2000)

    # Mudando o main frame do site que será navegado
    driver.switch_to.frame("main2")

    CodUasg = wdw.until(
        EC.presence_of_element_located((By.XPATH, "//input[@id='co_uasg']"))
    )
    CodUasg.send_keys("150182")

    NumPreg = wdw.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@id='numprp']")))
    NumPreg.send_keys(pregao)

    driver.execute_script("ValidaForm();")
    sleep(1)

    PrClick = wdw.until(
        EC.presence_of_element_located(
            (By.XPATH, "//a[contains(text(), '{}')]".format(pregao))
        )
    )
    PrClick.click()

    print("Preencha o CAPTCHA")

    wdw.until(
        EC.presence_of_element_located(
            (By.XPATH, "//center[contains(text(), 'MINISTÉRIO DA EDUCAÇÃO')]")
        )
    )
    print("Portal de compras encontrado")

    fail = ["Item deserto", "Cancelado no julgamento"]

    ############################## Alterei aqui ##############################
    if pd.isna(BaseDado.iloc[BaseDado.shape[0]-1,0]):
        qtd_itens = BaseDado.shape[0]-1
    else:
        qtd_itens = BaseDado.shape[0]

    item, collecting, page = [1, True, 0]

    while collecting:
        actual_item = (page)*100+item
        print(f'\n\n\n\n \U0001F3F8 Buscando pelo item: {actual_item}\n\n')

        Vencedores = wdw2.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//tr[contains(@class, 'tex3a')][{}]/td[6]".format(
                        item)
                ),
            )
        )

        if Vencedores.text in fail:
            print("Status do item ", actual_item, ":", Vencedores.text)
            BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc('Status')] = Vencedores.text
            if actual_item == qtd_itens:
                collecting = False
            elif item == 100:
                item -= 99
                page += 1
                driver.execute_script("javascript:PaginarItens('Proxima');")
            else:
                item += 1
            continue

        Vencedores = wdw2.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//tr[./td/a[contains(@href, 'javascript:void(0)')]][{}]/td[6]/a".format(
                        item),
                )
            )
        )

        print("Status do item ", actual_item, ":", Vencedores.text, "\n")
        BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc('Status')] = Vencedores.text
        Vencedores.click()
        driver.switch_to.window(driver.window_handles[1])
        Captcha = wdw2.until(
            EC.presence_of_element_located((By.XPATH, "//h2")))
        if Captcha.text == "ACOMPANHAMENTO DE LICITAÇÃO":
            print("Preencha o CAPTCHA")
            wdw.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(), 'Pregão/Concorrência Eletrônica')]")))

        vencedor = 1
        for colocado in range(1, 6):
            try:
                InfoEmpresa = []

                Cnpj = driver.find_element(By.XPATH,
                    "//tr[@class = 'tex3'][{}]/td[1]".format(colocado)
                )

                Nome = driver.find_element(By.XPATH,
                    "//tr[@class = 'tex3'][{}]/td[2]".format(colocado)
                )

                ValorEmpresa = driver.find_element(By.XPATH,
                    "//tr[@class = 'tex3'][{}]/td[4]".format(colocado)
                ).text

                StatusEmpresa = driver.find_element(By.XPATH,
                    "//tr[@class='tex3'][{}]/td[@class='tex3b']".format(
                        colocado)
                ).text

                ProdutoInfo = driver.find_element(By.XPATH,
                    "//tr[@class='tex5a'][{}]/td[@colspan='6']".format(
                        colocado)
                )
                ProdutoInfo = ProdutoInfo.text.split()

                Marca = []
                Fabricante = []
                Versão = []
                i = 0
                while ProdutoInfo[i] != "Descrição":
                    if ProdutoInfo[i] == "Marca:":
                        PosMarca = i
                    elif ProdutoInfo[i] == "Fabricante:":
                        PosFabricante = i
                    elif ProdutoInfo[i] == "Modelo":
                        PosVersão = i
                    i += 1

                for j in range(PosMarca + 1, PosFabricante):
                    Marca.append(ProdutoInfo[j])
                for j in range(PosFabricante + 1, PosVersão):
                    Fabricante.append(ProdutoInfo[j])
                for j in range(PosVersão + 3, i):
                    Versão.append(ProdutoInfo[j])

                Marca = " ".join(Marca)
                Fabricante = " ".join(Fabricante)
                Versão = " ".join(Versão)

                InfoEmpresa = [
                    ("CNPJ: " + Cnpj.text),
                    ("NOME: " + Nome.text),
                    ("R$ " + ValorEmpresa),
                    ("Marca: " + Marca),
                    ("Fabricante: " + Fabricante),
                    ("Modelo / Versão: " + Versão),
                ]

                BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')] = (
                    InfoEmpresa[0]
                    + "\n"
                    + InfoEmpresa[1]
                    + "\n"
                    + InfoEmpresa[2]
                    + "\n"
                    + InfoEmpresa[3]
                    + "\n"
                    + InfoEmpresa[4]
                    + "\n"
                    + InfoEmpresa[5]
                )

                print(InfoEmpresa)

                if StatusEmpresa == "Recusado":
                    vencedor += 1
                    BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')] = "Recusado \U0001F6AB \n" + \
                        BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')]
                    print(f"{StatusEmpresa} '\U0001F6AB'")
                elif StatusEmpresa == "Aceito":
                    BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')] = "Aceito \U00002705 \n" + \
                        BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')]
                    print(f"{StatusEmpresa} '\u2705'")
                elif StatusEmpresa == "Adjudicado":
                    BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')] = "Adjudicado \U00002705 \n" + \
                        BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')]
                    print(f"{StatusEmpresa} '\u2705'")

                if colocado == vencedor:
                    try:
                        ValorNegociado = driver.find_element(By.XPATH,
                            "//tr[@class='tex3'][{}]/td[not(node())]".format(colocado))
                        BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc('ValorFornecedor')] = float(
                            ValorEmpresa.replace(".", "").replace(",", "."))
                    except:
                        ValorNegociado = driver.find_element(By.XPATH,
                            "//tr[@class='tex3'][{}]/td[6]".format(colocado)).text
                        BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc('ValorFornecedor')] = float(
                            ValorNegociado.replace(".", "").replace(",", "."))
                        BaseDado.iloc[actual_item-1,
                                    BaseDado.columns.get_loc(f'Pos {colocado}')] = f"Valor Negociado: {ValorNegociado}\n"+BaseDado.iloc[actual_item-1, BaseDado.columns.get_loc(f'Pos {colocado}')]
            except:
                break

        driver.switch_to.window(driver.window_handles[0])
        driver.switch_to.frame("main2")
        if actual_item == qtd_itens:
            collecting = False
        elif item == 100:
            item -= 99
            page += 1
            driver.execute_script("javascript:PaginarItens('Proxima');")
        else:
            item += 1
        BaseDado.to_excel(excel_path, index=False)
    driver.quit()

    def apply_style(style, filtered_row_index_list, function, column):
        subset = pd.IndexSlice[filtered_row_index_list, column]
        style.map(func=functions[function],  subset=subset)

    functions = {
        'highlight_price': lambda x: 'background-color: red',
        'highlight_desert': lambda x: 'background-color: orange',
        'highlight_canceled': lambda x: 'background-color: #d65f5f',
        'highlight_refused': lambda x: 'background-color: #8147d1',
        'highlight_accepted': lambda x: 'background-color: #5cde35',
        'highlight_awarded': lambda x: 'background-color: #eb38af'
    }

    status_list = ['Recusado', 'Aceito', 'Adjudicado']

    positions = ['Pos 1', 'Pos 2', 'Pos 3', 'Pos 4', 'Pos 5']

    style = BaseDado.style

    style.map(
            functions['highlight_price'], subset=pd.IndexSlice[list(BaseDado.query(
                "`VALOR DE REFERÊNCIA (unitário)(R$)` < `ValorFornecedor`").index), 'ValorFornecedor']
        )

    style.map(
            functions['highlight_desert'], subset=pd.IndexSlice[list(
                BaseDado[BaseDado['Status'] == 'Item deserto'].index), 'Status']
        )

    style.map(
            functions['highlight_canceled'], subset=pd.IndexSlice[list(
                BaseDado[BaseDado['Status'] == 'Cancelado no julgamento'].index), 'Status']
        )

    for tupla in zip(list(functions.keys())[3:], status_list):
        function, status = tupla
        for position in positions:
            apply_style(
                style=style,
                filtered_row_index_list=list(BaseDado[BaseDado[position].str.contains(status, na=False)].index),
                function=function,
                column=position
            )

    style.to_excel(excel_path, index=False)

    print('\nExtração Concluída \U0001F7E2')
    print('\nIniciando a estruturação da planilha... \U000023F3')

    ####### INÍCIO DO OPENPYXL ########

    wb = load_workbook(filename = rf'{excel_path}')

    ws = wb.active

    ws['A1'] = "ITEM"
    ws['B1'] = "DESCRIÇÃO / ESPECIFICAÇÃO"
    ws['C1'] = "SUGESTÃO DE CATMAT"
    ws['D1'] = "UNIDADE DE MEDIDA"
    ws['E1'] = "QUANTIDADE TOTAL"
    ws['F1'] = "VALOR DE REFERÊNCIA (unitário) (R$)"
    ws['G1'] = "VALOR FIRMA VENCEDORA"
    ws['H1'] = "POSIÇÃO: 1"
    ws['I1'] = "POSIÇÃO: 2"
    ws['J1'] = "POSIÇÃO: 3"
    ws['K1'] = "POSIÇÃO: 4"
    ws['L1'] = "POSIÇÃO: 5"
    ws['M1'] = "STATUS DO ITEM"

    ws.sheet_view.zoomScale = 70

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # ao inserir, tratar as colunas como números iniciando em 1
    def modify(index, rows_length, ByRow=False, hOrientation='left', vOrientation='top', wrap_text=False, row_space=None, col_space=None, currency=False, border=False, 
        font_name='Calibri', font_size='10', font_color='000000', font_italic=False, font_bold=False, cell_color=None):
        if type(index)==int:
            index = [index]
        column_count = ws.max_column
        for vector in index:
            if ByRow:
                for col in range(1, column_count+1):
                    currentCell = ws[f'{get_column_letter(col)}{vector}']
                    currentCell.alignment = Alignment(horizontal = hOrientation, vertical = vOrientation, wrapText=wrap_text)
                    if border==True:
                        currentCell.border = thin_border
                    ft = Font(name=font_name, size=font_size, color=font_color, italic=font_italic, bold=font_bold)
                    currentCell.font = ft
                    if cell_color is not None:
                        currentCell.fill = PatternFill("solid", start_color=cell_color)
                    if col_space is not None:
                        letter = get_column_letter(col)
                        ws.column_dimensions[letter].width = col_space
                if row_space is not None:
                    ws.row_dimensions[vector].height = row_space
            else:
                for row in range(2, rows_length+2):
                    currentCell = ws[f'{get_column_letter(vector)}{row}']
                    currentCell.alignment = Alignment(horizontal = hOrientation, vertical = vOrientation, wrapText=wrap_text)
                    if currency==True:
                        currentCell.number_format = 'R$ #,##0.00'
                    if border==True:
                        currentCell.border = thin_border
                    ft = Font(name=font_name, size=font_size, color=font_color, italic=font_italic, bold=font_bold)
                    currentCell.font = ft
                    if cell_color is not None:
                        currentCell.fill = PatternFill("solid", start_color=cell_color)
                    if row_space is not None:
                        ws.row_dimensions[row].height = row_space
                if col_space is not None:
                    letter = get_column_letter(vector)
                    ws.column_dimensions[letter].width = col_space
                

    modify(index=1, rows_length=qtd_itens, ByRow=True, hOrientation='center', vOrientation='center', wrap_text=True, font_bold=True, row_space=60, cell_color='f5425a')
    modify(index=1, rows_length=qtd_itens, hOrientation='center', vOrientation='center', cell_color='A5B0A2', font_bold=True, row_space=140, col_space=7, border=True)
    modify(index=2, rows_length=qtd_itens, border=True, wrap_text=True, col_space=34)
    modify(index=[3, 4, 5], rows_length=qtd_itens, border=True, col_space=14, hOrientation='center', vOrientation='center')
    modify(index=[6, 7], rows_length=qtd_itens, border=True, col_space=14, hOrientation='center', vOrientation='center', currency=True)
    modify(index=[8,9,10,11,12], rows_length=qtd_itens, border=True, col_space=30, hOrientation='center', vOrientation='center', wrap_text=True)
    modify(index=13, rows_length=qtd_itens, border=True, col_space=22, hOrientation='center', vOrientation='center')

    wb.save(excel_path)

    print('\nEstruturação Concluida! \U0001F47E')